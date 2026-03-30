# import pandas as pd

# def build_excel(sections, fetch_func, file_path):
#     with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
#         for section_id, section_name in sections:
#             rows = fetch_func(section_id)

#             if not rows:
#                 continue

#             df = pd.DataFrame(rows, columns=["Serial", "Name", "Date", "Status"])

#             # Pivot: rows -> students, columns -> dates
#             pivot = df.pivot_table(
#                 index=["Serial", "Name"],
#                 columns="Date",
#                 values="Status",
#                 aggfunc="first"
#             ).reset_index()

#             pivot.to_excel(writer, sheet_name=section_name, index=False)

import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font

def build_excel(sections, fetch_func, file_path):
    # 1. Write basic data sheets
    with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
        for section_id, section_name in sections:
            rows = fetch_func(section_id)
            if not rows:
                # Create empty sheet if no data
                pd.DataFrame().to_excel(writer, sheet_name=section_name)
                continue

            df = pd.DataFrame(rows, columns=["Serial", "Name", "Date", "Status"])
            df["Date"] = pd.to_datetime(df["Date"])

            pivot = df.pivot_table(
                index=["Serial", "Name"],
                columns="Date",
                values="Status",
                aggfunc="first"
            ).reset_index()

            # Format columns as YYYY-MM-DD for consistency
            pivot.columns = [
                c.strftime("%Y-%m-%d") if isinstance(c, datetime) else c 
                for c in pivot.columns
            ]

            pivot.to_excel(writer, sheet_name=section_name, index=False)

    # 2. Post-process to add Month -> Week headers
    wb = load_workbook(file_path)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if ws.max_column < 3: # serial, name, at least one date
            continue
            
        # Insert 2 rows for headers (Month, Week)
        ws.insert_rows(1, amount=2)
        
        # Row 3 is now the date headers (formerly Row 1)
        # Parse dates from Row 3 (cols 3+)
        date_cols = []
        for col in range(3, ws.max_column + 1):
            cell_val = ws.cell(row=3, column=col).value
            if cell_val:
                try:
                    dt = datetime.strptime(str(cell_val), "%Y-%m-%d").date()
                    date_cols.append((col, dt))
                except ValueError:
                    pass

        if not date_cols:
            continue

        # Group by Month -> Week
        # month_key: (year, month) -> { week_key: (year, week) -> [cols] }
        structure = {}
        
        for col, dt in date_cols:
            m_key = (dt.year, dt.month)
            w_key = dt.isocalendar()[:2] # (year, week)
            
            if m_key not in structure:
                structure[m_key] = {}
            if w_key not in structure[m_key]:
                structure[m_key][w_key] = []
            
            structure[m_key][w_key].append(col)
            
        # Draw Headers
        # Month on Row 1
        # Week on Row 2
        
        font_bold = Font(bold=True)
        center_align = Alignment(horizontal='center')

        for (year, month), weeks in structure.items():
            month_name = datetime(year, month, 1).strftime("%B %Y")
            
            # Find min and max col for this month
            all_cols = [c for w_cols in weeks.values() for c in w_cols]
            min_c, max_c = min(all_cols), max(all_cols)
            
            # Merge Month
            ws.merge_cells(start_row=1, start_column=min_c, end_row=1, end_column=max_c)
            c = ws.cell(row=1, column=min_c)
            c.value = month_name
            c.font = font_bold
            c.alignment = center_align
            
            # Merge Weeks
            for (w_year, w_num), cols in weeks.items():
                w_min, w_max = min(cols), max(cols)
                ws.merge_cells(start_row=2, start_column=w_min, end_row=2, end_column=w_max)
                cw = ws.cell(row=2, column=w_min)
                cw.value = f"Week {w_num}"
                cw.font = font_bold
                cw.alignment = center_align

        # Fix original headers (Serial, Name) position
        # They are at Row 3, Col 1 & 2. 
        # Clean up empty cells above them
        ws.cell(row=3, column=1).value = "Serial"
        ws.cell(row=3, column=2).value = "Name"
        
    wb.save(file_path)
